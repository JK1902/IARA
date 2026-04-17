"""
Flask Backend Integration for Enhanced Student Scoring System
Integrates the enhanced essay analyzer with your existing frontend
"""
from dotenv import load_dotenv
load_dotenv()
from openai import OpenAI
from services.financial_fraud_detector import analyze_financial_pdf
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from services.student_analyzer import StudentAnalyzerSafe
from db.database import get_db
from db.report_generator import generate_report as build_pdf_report
from services.nlp_service import NLPService
import io
import csv
import os
import re
import uuid
from typing import Optional, Dict
from textwrap import shorten
from werkzeug.utils import secure_filename
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import simpleSplit
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from services.financial_fraud_detector import analyze_financial_pdf
from services.transcript_fraud_detector import analyze_transcript_pdf

nlp_service = NLPService()
app = Flask(__name__)
app.secret_key = 'your-secret-key-here-change-in-production'

# Initialize the enhanced analyzer
analyzer = StudentAnalyzerSafe()
analysis_results = []

# --- SAMPLE DATA ---
SAMPLE_DATA = {
    'high': {
        'studentId': 'STU_HIGH_001',
        'country': 'India',
        'gpa': 3.9,
        'curriculum': 'US HS/University',
        'travelHistory': 'SEVIS/Multiple US trips',
        'essayText': 'Throughout my academic journey, I have consistently demonstrated a passion for engineering and innovation. My research experience at IIT Delhi, combined with my internship at a leading tech firm, has prepared me well for graduate study. I am particularly interested in machine learning applications in healthcare, and I believe WSU\'s program aligns perfectly with my goals.',
        'negFactors': []
    },
    'medium': {
        'studentId': 'STU_MED_001',
        'country': 'Vietnam',
        'gpa': 3.2,
        'curriculum': 'IGCSE/IB',
        'travelHistory': 'Multiple listed',
        'essayText': 'I have always been passionate about computer science. During my undergraduate studies, I worked on several projects involving data analysis and software development. I hope to further my education at WSU and contribute to the research community.',
        'negFactors': ['bankDocsPending']
    },
    'low': {
        'studentId': 'STU_LOW_001',
        'country': 'Nigeria',
        'gpa': 2.4,
        'curriculum': 'Standard Intl Secondary',
        'travelHistory': 'No travel abroad',
        'essayText': 'I want to study at WSU because it is a good university. I will work hard and do my best.',
        'negFactors': ['reqAppFeeWaiver', 'cannotPayFee', 'bankDocsPending']
    }
}

# --- ROUTES ---
@app.route('/')
def index():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('index.html', username=session.get('username', 'User'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json(silent=True) or request.form
        username = data.get("username")
        password = data.get("password")
        if not username or not password:
            return jsonify({"success": False, "error": "Username and password required"})
        db = get_db()
        user = db.authenticate_user(username, password)
        if user:
            session["user_id"] = user['id']
            session["username"] = user['username']
            session["full_name"] = user.get('full_name') or user['username']
            session["role"] = user['role']
            return jsonify({"success": True, "redirect": url_for("index")})
        else:
            return jsonify({"success": False, "error": "Invalid credentials"})
    return render_template("login.html")

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

def require_login():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return None

@app.route('/api/sample/<type>')
def get_sample(type):
    if type not in SAMPLE_DATA:
        return jsonify({'error': 'Invalid sample type'}), 400
    return jsonify(SAMPLE_DATA[type])

# --- STUDENT ANALYSIS ---
@app.route("/api/analyze", methods=["POST"])
def analyze_student():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No input data provided'}), 400

        essay = data.get("essayText") or data.get("essay")
        prompt = data.get("prompt", "")
        if not essay:
            return jsonify({'error': 'No essay text provided'}), 400

        sentiment = nlp_service.analyze_sentiment(essay)
        similarity = nlp_service.compute_similarity(essay, prompt)

        gpa = float(data.get('gpa', 0))
        curriculum = data.get('curriculum', '')
        travel_history = data.get('travelHistory', '')
        neg_factors = data.get('negFactors', [])

        result_resp = analyzer.analyze_student_safe(
            gpa=gpa,
            curriculum=curriculum,
            travel_history=travel_history,
            essay_text=essay,
            neg_factors=neg_factors
        )

        # Convert to dict safely
        if isinstance(result_resp, dict):
            result = result_resp
        elif hasattr(result_resp, "model_dump"):
            result = result_resp.model_dump()
        elif hasattr(result_resp, "dict") and callable(getattr(result_resp, "dict", None)):
            result = result_resp.dict()
        else:
            try:
                result = dict(result_resp)
            except:
                result = result_resp

        response = {
            'posScore': float(result.get('pos_score', 0)),
            'negScore': float(result.get('neg_score', 0)),
            'finalScore': float(result.get('final_score', 0)),
            'breakdown': result.get('breakdown', {}),
            'rankEstimate': result.get('rank_estimate', "N/A"),
            'recommendation': result.get('recommendation', "No recommendation"),
            'essayAnalysis': {
                'clarity_focus': result.get('essay_analysis', {}).get('clarity_focus', ''),
                'development_organization': result.get('essay_analysis', {}).get('development_organization', ''),
                'creativity_style': result.get('essay_analysis', {}).get('creativity_style', ''),
                'rubric_score': float(result.get('essay_analysis', {}).get('rubric_score', 0)),
                'grammar_score': float(result.get('essay_analysis', {}).get('grammar_score', 0)),
                'coherence_score': float(result.get('essay_analysis', {}).get('coherence_score', 0)),
                'vocabulary_richness': float(result.get('essay_analysis', {}).get('vocabulary_richness', 0)),
                'insights': result.get('essay_analysis', {}).get('insights', []),
            },
            'overall_confidence': float(result.get('overall_confidence', 0))
        }

        # ── Save to in-memory results so analytics dashboard is populated ──
        analysis_results.append({
            'timestamp': datetime.now().isoformat(),
            'studentId': data.get('studentId', 'N/A'),
            'country': data.get('country', 'N/A'),
            'gpa': gpa,
            'curriculum': curriculum,
            'travelHistory': travel_history,
            'essayLength': len(essay),
            'negFactors': ', '.join(neg_factors),
            'posScore': response['posScore'],
            'negScore': response['negScore'],
            'finalScore': response['finalScore'],
            'rankEstimate': response['rankEstimate'],
            'recommendation': response['recommendation'],
            'clarityFocus': response['essayAnalysis'].get('clarity_focus', ''),
            'developmentOrg': response['essayAnalysis'].get('development_organization', ''),
            'creativityStyle': response['essayAnalysis'].get('creativity_style', ''),
            'essayRubricScore': response['essayAnalysis'].get('rubric_score', 0),
            'grammarScore': response['essayAnalysis'].get('grammar_score', 0),
            'coherenceScore': response['essayAnalysis'].get('coherence_score', 0),
            'vocabularyRichness': response['essayAnalysis'].get('vocabulary_richness', 0),
            'analysisConfidence': response.get('overall_confidence', 0)
        })

        return jsonify(response)

    except Exception as e:
        print(f"Analysis error: {e}")
        return jsonify({'error': str(e)}), 500

# --- BATCH ANALYSIS ---
@app.route('/api/batch-analyze', methods=['POST'])
def batch_analyze():
    try:
        data = request.get_json()
        students = data.get('students', [])
        if not students:
            return jsonify({'error': 'No students provided'}), 400

        batch_results = []

        for student_data in students:
            try:
                student_id = student_data.get('studentId', 'N/A')
                country = student_data.get('country', 'N/A')
                gpa = float(student_data.get('gpa', 0))
                curriculum = student_data.get('curriculum', '')
                travel_history = student_data.get('travelHistory', '')
                essay_text = student_data.get('essayText', '')
                neg_factors = student_data.get('negFactors', [])

                if not curriculum or not travel_history:
                    batch_results.append({
                        'studentId': student_id,
                        'success': False,
                        'error': 'Missing required fields'
                    })
                    continue

                result_resp = analyzer.analyze_student_safe(
                    gpa=gpa,
                    curriculum=curriculum,
                    travel_history=travel_history,
                    essay_text=essay_text,
                    neg_factors=neg_factors
                )

                # Convert to dict safely
                if isinstance(result_resp, dict):
                    result = result_resp
                elif hasattr(result_resp, "model_dump"):
                    result = result_resp.model_dump()
                elif hasattr(result_resp, "dict") and callable(getattr(result_resp, "dict", None)):
                    result = result_resp.dict()
                else:
                    try:
                        result = dict(result_resp)
                    except:
                        result = result_resp

                ea = result.get('essay_analysis', {})

                analysis_record = {
                    'timestamp': datetime.now().isoformat(),
                    'studentId': student_id,
                    'country': country,
                    'gpa': gpa,
                    'curriculum': curriculum,
                    'travelHistory': travel_history,
                    'essayLength': len(essay_text),
                    'negFactors': ', '.join(neg_factors),
                    'posScore': float(result.get('pos_score', 0)),
                    'negScore': float(result.get('neg_score', 0)),
                    'finalScore': float(result.get('final_score', 0)),
                    'rankEstimate': result.get('rank_estimate', 'N/A'),
                    'recommendation': result.get('recommendation', 'No recommendation'),
                    'clarityFocus': ea.get('clarity_focus', ''),
                    'developmentOrg': ea.get('development_organization', ''),
                    'creativityStyle': ea.get('creativity_style', ''),
                    'essayRubricScore': float(ea.get('rubric_score', 0)),
                    'grammarScore': float(ea.get('grammar_score', 0)),
                    'coherenceScore': float(ea.get('coherence_score', 0)),
                    'vocabularyRichness': float(ea.get('vocabulary_richness', 0)),
                    'analysisConfidence': float(result.get('overall_confidence', 0))
                }

                analysis_results.append(analysis_record)

                batch_results.append({
                    'studentId': student_id,
                    'country': country,
                    'success': True,
                    'posScore': analysis_record['posScore'],
                    'negScore': analysis_record['negScore'],
                    'finalScore': analysis_record['finalScore'],
                    'rankEstimate': analysis_record['rankEstimate'],
                    'recommendation': analysis_record['recommendation']
                })

            except Exception as e:
                batch_results.append({
                    'studentId': student_data.get('studentId', 'Unknown'),
                    'success': False,
                    'error': str(e)
                })

        return jsonify({'results': batch_results})

    except Exception as e:
        print(f"Batch analysis error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/csv', methods=['POST'])
def export_csv():
    try:
        if not analysis_results:
            return jsonify({'error': 'No data to export'}), 400

        output = io.StringIO()
        headers = [
            'Timestamp', 'Student ID', 'Country', 'GPA', 'Curriculum',
            'Travel History', 'Essay Length', 'Neg Factors',
            'POS Score', 'NEG Score', 'Final Score', 'Rank Estimate',
            'Clarity & Focus', 'Development & Org', 'Creativity & Style',
            'Essay Rubric Score', 'Grammar Score', 'Coherence Score',
            'Vocabulary Richness', 'Analysis Confidence', 'Recommendation'
        ]

        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()

        for record in analysis_results:
            writer.writerow({
                'Timestamp': record['timestamp'],
                'Student ID': record['studentId'],
                'Country': record['country'],
                'GPA': record['gpa'],
                'Curriculum': record['curriculum'],
                'Travel History': record['travelHistory'],
                'Essay Length': record['essayLength'],
                'Neg Factors': record['negFactors'],
                'POS Score': record['posScore'],
                'NEG Score': record['negScore'],
                'Final Score': record['finalScore'],
                'Rank Estimate': record['rankEstimate'],
                'Clarity & Focus': record['clarityFocus'],
                'Development & Org': record['developmentOrg'],
                'Creativity & Style': record['creativityStyle'],
                'Essay Rubric Score': record['essayRubricScore'],
                'Grammar Score': record['grammarScore'],
                'Coherence Score': record['coherenceScore'],
                'Vocabulary Richness': record['vocabularyRichness'],
                'Analysis Confidence': record['analysisConfidence'],
                'Recommendation': record['recommendation']
            })

        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'student_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )

    except Exception as e:
        print(f"CSV export error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    try:
        if not analysis_results:
            return jsonify({'error': 'No data to export'}), 400

        wb = Workbook()
        ws = wb.active
        ws.title = "Student Analysis"

        headers = [
            'Timestamp', 'Student ID', 'Country', 'GPA', 'Curriculum',
            'Travel History', 'Essay Length', 'Neg Factors',
            'POS Score', 'NEG Score', 'Final Score', 'Rank Estimate',
            'Clarity & Focus', 'Development & Org', 'Creativity & Style',
            'Essay Rubric Score', 'Grammar Score', 'Coherence Score',
            'Vocabulary Richness', 'Analysis Confidence', 'Recommendation'
        ]

        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_idx, record in enumerate(analysis_results, start=2):
            ws.cell(row=row_idx, column=1,  value=record['timestamp'])
            ws.cell(row=row_idx, column=2,  value=record['studentId'])
            ws.cell(row=row_idx, column=3,  value=record['country'])
            ws.cell(row=row_idx, column=4,  value=record['gpa'])
            ws.cell(row=row_idx, column=5,  value=record['curriculum'])
            ws.cell(row=row_idx, column=6,  value=record['travelHistory'])
            ws.cell(row=row_idx, column=7,  value=record['essayLength'])
            ws.cell(row=row_idx, column=8,  value=record['negFactors'])
            ws.cell(row=row_idx, column=9,  value=record['posScore'])
            ws.cell(row=row_idx, column=10, value=record['negScore'])
            ws.cell(row=row_idx, column=11, value=record['finalScore'])
            ws.cell(row=row_idx, column=12, value=record['rankEstimate'])
            ws.cell(row=row_idx, column=13, value=record['clarityFocus'])
            ws.cell(row=row_idx, column=14, value=record['developmentOrg'])
            ws.cell(row=row_idx, column=15, value=record['creativityStyle'])
            ws.cell(row=row_idx, column=16, value=record['essayRubricScore'])
            ws.cell(row=row_idx, column=17, value=record['grammarScore'])
            ws.cell(row=row_idx, column=18, value=record['coherenceScore'])
            ws.cell(row=row_idx, column=19, value=record['vocabularyRichness'])
            ws.cell(row=row_idx, column=20, value=record['analysisConfidence'])
            ws.cell(row=row_idx, column=21, value=record['recommendation'])

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'student_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        print(f"Excel export error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/generate-report', methods=['POST'])
def generate_report_endpoint():
    """Generate report and save to database."""
    try:
        data           = request.get_json()
        student_data   = data.get('studentData', {})
        result_data    = data.get('resultData', {})
        staff_comments = data.get('staffComments', '')
        reviewer_name  = data.get('reviewerName', session.get('full_name', 'Staff'))
        fmt            = data.get('format', 'pdf').lower()
        analysis_id    = data.get('analysisId')

        student_id = student_data.get('studentId', 'N/A')
        db = get_db()

        if fmt == 'pdf':
            pdf_buffer = build_pdf_report(
                student_data   = student_data,
                result_data    = result_data,
                staff_comments = staff_comments,
                reviewer_name  = reviewer_name,
            )
            pdf_bytes = pdf_buffer.read()

            # Save to database using the actual signature: (analysis_id, student_id, user_id, pdf_blob)
            db.save_report(
                analysis_id = analysis_id,
                student_id  = student_id,
                user_id     = session.get('user_id', 0),
                pdf_blob    = pdf_bytes
            )

            return send_file(
                io.BytesIO(pdf_bytes),
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f"admission_report_{student_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
            )

        else:  # txt
            now  = datetime.now().strftime('%B %d, %Y  %I:%M %p')
            sep  = '=' * 65
            thin = '-' * 65

            neg_str = ', '.join(student_data.get('negFactors', [])) or 'None'

            lines = [
                sep,
                'WSU OFFICE OF INTERNATIONAL PROGRAMS',
                'ADMISSION ANALYSIS REPORT — CONFIDENTIAL',
                sep,
                f"Student ID   : {student_id}",
                f"Country      : {student_data.get('country', 'N/A')}",
                f"Reviewed by  : {reviewer_name}",
                f"Report Date  : {now}",
                sep,
                '',
                'SCORE SUMMARY',
                thin,
                f"  POS Score    : +{float(result_data.get('posScore', 0)):.2f}",
                f"  NEG Score    :  -{abs(float(result_data.get('negScore', 0))):.2f}",
                f"  FINAL Score  :  {float(result_data.get('finalScore', 0)):.2f}",
                '',
                'RECOMMENDATION',
                thin,
                f"  {result_data.get('recommendation', 'N/A')}",
                '',
                'STAFF COMMENTS',
                thin,
                staff_comments.strip() if staff_comments.strip() else '(No additional comments recorded.)',
                '',
                sep,
            ]

            txt_content = '\n'.join(lines)

            # Save to database using the actual signature: (analysis_id, student_id, user_id, pdf_blob)
            db.save_report(
                analysis_id = analysis_id,
                student_id  = student_id,
                user_id     = session.get('user_id', 0),
                pdf_blob    = txt_content.encode('utf-8')
            )

            return send_file(
                io.BytesIO(txt_content.encode('utf-8')),
                mimetype='text/plain',
                as_attachment=True,
                download_name=f"admission_report_{student_id}_{datetime.now().strftime('%Y%m%d')}.txt"
            )

    except Exception as e:
        print(f"Report generation error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/reports')
def reports_page():
    redirect_response = require_login()
    if redirect_response:
        return redirect_response
    db = get_db()
    reports = db.get_all_reports(limit=100) if hasattr(db, 'get_all_reports') else []
    return render_template('report.html',
                           username=session.get('full_name'),
                           reports=reports)


@app.route('/student/<student_id>')
def student_detail(student_id):
    redirect_response = require_login()
    if redirect_response:
        return redirect_response
    db = get_db()
    analyses = db.get_student_analyses(student_id) if hasattr(db, 'get_student_analyses') else []
    reports  = db.get_student_reports(student_id)  if hasattr(db, 'get_student_reports')  else []
    return render_template('student_detail.html',
                           username=session.get('full_name'),
                           student_id=student_id,
                           analyses=analyses,
                           reports=reports)


@app.route('/api/report/<int:report_id>/download')
def download_report(report_id):
    redirect_response = require_login()
    if redirect_response:
        return redirect_response
    db = get_db()
    report = db.get_report_by_id(report_id) if hasattr(db, 'get_report_by_id') else None
    if not report:
        return jsonify({'error': 'Report not found'}), 404
    return send_file(
        io.BytesIO(report['pdf_blob']),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f"report_{report['student_id']}_{report_id}.pdf"
    )

@app.route('/admin/users')
def admin_users():
    redirect_response = require_login()
    if redirect_response:
        return redirect_response
    if session.get('role') != 'admin':
        return "Access denied", 403
    db = get_db()
    users    = db.get_all_users()    if hasattr(db, 'get_all_users')    else []
    activity = db.get_activity_log(limit=50)
    return render_template('admin_users.html',
                           username=session.get('full_name'),
                           users=users,
                           activity=activity)


@app.route('/api/admin/create-user', methods=['POST'])
def create_user():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Access denied'}), 403
    data      = request.get_json()
    username  = data.get('username')
    password  = data.get('password')
    full_name = data.get('full_name')
    email     = data.get('email')
    role      = data.get('role', 'reviewer')
    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400
    db      = get_db()
    user_id = db.create_user(username, password)
    if user_id:
        return jsonify({'success': True, 'user_id': user_id})
    else:
        return jsonify({'error': 'Username already exists'}), 400

@app.route('/api/admin/toggle-user', methods=['POST'])
def toggle_user():
    """Toggle user active/inactive status."""
    if session.get('role') != 'admin':
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        is_active = data.get('is_active')  # Will be 0 or 1
        
        if not user_id:
            return jsonify({'error': 'User ID required'}), 400
        
        db = get_db()
        
        # Update user status in database
        if hasattr(db, 'toggle_user_status'):
            success = db.toggle_user_status(user_id, is_active)
        else:
            # Fallback if method doesn't exist
            return jsonify({'error': 'Database method not implemented'}), 500
        
        if success:
            # Log the activity
            action = 'user_deactivated' if is_active == 0 else 'user_activated'
            if hasattr(db, 'log_activity'):
                db.log_activity(
                    user_id=session.get('user_id'),
                    action=action,
                    details=f"User ID {user_id}"
                )
            
            return jsonify({'success': True, 'message': 'User status updated'})
        else:
            return jsonify({'error': 'Failed to update user'}), 400
            
    except Exception as e:
        print(f"Toggle user error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/delete-user', methods=['POST'])
def delete_user():
    """Permanently delete a user."""
    if session.get('role') != 'admin':
        return jsonify({'error': 'Access denied'}), 403
    
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'User ID required'}), 400
        
        db = get_db()
        
        # Prevent deleting the admin user
        user = db.get_user_by_id(user_id) if hasattr(db, 'get_user_by_id') else None
        if user and user.get('username') == 'admin':
            return jsonify({'error': 'Cannot delete admin user'}), 403
        
        # Delete user from database
        if hasattr(db, 'delete_user'):
            success = db.delete_user(user_id)
        else:
            return jsonify({'error': 'Database method not implemented'}), 500
        
        if success:
            # Log the activity
            if hasattr(db, 'log_activity'):
                db.log_activity(
                    user_id=session.get('user_id'),
                    action='user_deleted',
                    details=f"Deleted user ID {user_id}"
                )
            
            return jsonify({'success': True, 'message': 'User deleted'})
        else:
            return jsonify({'error': 'Failed to delete user'}), 400
            
    except Exception as e:
        print(f"Delete user error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/dashboard')
def dashboard():
    redirect_response = require_login()
    if redirect_response:
        return redirect_response
    db = get_db()
    stats          = db.get_dashboard_stats()   if hasattr(db, 'get_dashboard_stats') else {}
    recent_reports = db.get_all_reports(limit=10) if hasattr(db, 'get_all_reports')   else []
    return render_template('dashboard.html',
                           username=session.get('full_name'),
                           stats=stats,
                           recent_reports=recent_reports)


def _build_analytics_data():
    """Shared helper used by both /analytics and /api/analytics/summary."""
    total_analyzed    = len(analysis_results)
    distribution      = {'highPotential': 0, 'mediumRisk': 0, 'highRisk': 0}
    country_counts    = {}
    total_final_score = 0
    total_grammar_score = 0

    for r in analysis_results:
        final_score = r.get('finalScore', 0)
        total_final_score   += final_score
        total_grammar_score += r.get('grammarScore', 0)

        recommendation = r.get('recommendation', '').lower()
        if 'highly recommended' in recommendation or final_score >= 7.5:
            distribution['highPotential'] += 1
        elif 'recommended' in recommendation or final_score >= 6.5:
            distribution['mediumRisk'] += 1
        else:
            distribution['highRisk'] += 1

        country = r.get('country', 'Unknown')
        country_counts[country] = country_counts.get(country, 0) + 1

    avg_final_score   = round(total_final_score   / total_analyzed, 2) if total_analyzed else 0
    avg_grammar_score = round(total_grammar_score / total_analyzed, 2) if total_analyzed else 0

    top_countries = sorted(
        [{'country': k, 'count': v} for k, v in country_counts.items()],
        key=lambda x: x['count'], reverse=True
    )[:10]

    country_scores = {}
    for r in analysis_results:
        country = r.get('country', 'Unknown')
        country_scores.setdefault(country, []).append(r.get('finalScore', 0))
    avg_scores_per_country = {
        c: round(sum(scores) / len(scores), 2) for c, scores in country_scores.items()
    }

    return {
        'totalStudents':       total_analyzed,
        'avgFinalScore':       avg_final_score,
        'avgGrammarScore':     avg_grammar_score,
        'distribution':        distribution,
        'countryStats':        top_countries,
        'avgScoresPerCountry': avg_scores_per_country,
        'results':             analysis_results[-10:]
    }


@app.route('/analytics')
def analytics():
    if 'username' not in session:
        return redirect(url_for('login'))
    analytics_data = _build_analytics_data()
    return render_template('analytics.html',
                           username=session.get('username'),
                           analytics=analytics_data)


@app.route('/api/analytics/summary')
def analytics_summary():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    data = _build_analytics_data()
    data.pop('results', None)   # don't send full result list in summary API
    return jsonify(data)


@app.route('/batch')
def batch():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('batch.html', username=session.get('username'))


@app.route('/financial')
def financial():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('financial.html', username=session.get('username'))


@app.route('/transcript')
def transcript():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('transcript.html', username=session.get('username'))

# ==================== FRAUD DETECTION API ====================

@app.route('/api/fraud/financial', methods=['POST'])
def analyze_financial_document():
    """Upload a financial PDF and run the fraud detector."""
    redirect_response = require_login()
    if redirect_response:
        return redirect_response

    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file part in request'}), 400

        file = request.files['file']
        if not file or file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        upload_dir = os.path.join(os.path.dirname(__file__), 'tmp_uploads')
        os.makedirs(upload_dir, exist_ok=True)

        safe_name = secure_filename(file.filename or "uploaded_financial.pdf")
        unique = uuid.uuid4().hex
        temp_path = os.path.join(upload_dir, f"{unique}_{safe_name}")

        file.save(temp_path)

        try:
            result = analyze_financial_pdf(temp_path, max_pages=20)
            return jsonify(result)
        finally:
            try:
                os.remove(temp_path)
            except OSError:
                pass

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/fraud/financial/report', methods=['POST'])
def export_financial_fraud_report():
    """Export one financial fraud report as PDF or TXT."""
    redirect_response = require_login()
    if redirect_response:
        return redirect_response

    try:
        payload = request.get_json(silent=True) or {}
        result = payload.get("result")

        if not isinstance(result, dict):
            return jsonify({"error": "Missing or invalid 'result' object"}), 400

        comments = payload.get("comments", "") or ""
        meta = payload.get("meta", {}) or {}
        fmt = (payload.get("format", "pdf") or "pdf").lower()
        original_filename = payload.get("originalFilename", "") or ""

        reviewer = session.get("full_name") or session.get("username") or "unknown"

        report_text = build_financial_txt_report(
            result=result,
            reviewer=reviewer,
            comments=comments,
            meta=meta
        )

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = meta.get("studentId") or os.path.splitext(original_filename)[0] or "financial_fraud"
        safe_base = re.sub(r"[^A-Za-z0-9_\-]+", "_", base).strip("_") or "financial_fraud"

        if fmt == "txt":
            return send_file(
                io.BytesIO(report_text.encode("utf-8")),
                mimetype="text/plain",
                as_attachment=True,
                download_name=f"{safe_base}_report_{ts}.txt"
            )

        pdf_bytes = build_pdf_bytes_from_text(report_text)
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"{safe_base}_report_{ts}.pdf"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/fraud/financial/report/all', methods=['POST'])
def export_financial_fraud_report_all():
    """Export one combined financial fraud report for multiple documents."""
    redirect_response = require_login()
    if redirect_response:
        return redirect_response

    try:
        payload = request.get_json(silent=True) or {}
        documents = payload.get("documents", [])
        fmt = (payload.get("format", "pdf") or "pdf").lower()

        if not isinstance(documents, list) or len(documents) == 0:
            return jsonify({"error": "No documents provided for export."}), 400

        cleaned = []
        for d in documents:
            if not isinstance(d, dict):
                continue
            r = d.get("result")
            if not isinstance(r, dict):
                continue
            cleaned.append({
                "originalFilename": d.get("originalFilename", "") or "",
                "result": r,
                "comments": d.get("comments", "") or "",
                "meta": d.get("meta", {}) or {}
            })

        if len(cleaned) == 0:
            return jsonify({"error": "No valid document results found for export."}), 400

        reviewer = session.get("full_name") or session.get("username") or "unknown"
        report_text = build_multi_financial_txt_report(cleaned, reviewer=reviewer)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_base = "financial_fraud_combined"

        if fmt == "txt":
            return send_file(
                io.BytesIO(report_text.encode("utf-8")),
                mimetype="text/plain",
                as_attachment=True,
                download_name=f"{safe_base}_{ts}.txt"
            )

        pdf_bytes = build_pdf_bytes_from_text(report_text)
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"{safe_base}_{ts}.pdf"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/fraud/transcript', methods=['POST'])
def analyze_transcript_document():
    """Upload a transcript PDF and run the fraud detector."""
    redirect_response = require_login()
    if redirect_response:
        return redirect_response

    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file part in request'}), 400

        file = request.files['file']
        if not file or file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        upload_dir = os.path.join(os.path.dirname(__file__), 'tmp_uploads')
        os.makedirs(upload_dir, exist_ok=True)

        safe_name = secure_filename(file.filename or "uploaded_transcript.pdf")
        unique = uuid.uuid4().hex
        temp_path = os.path.join(upload_dir, f"{unique}_{safe_name}")

        file.save(temp_path)

        try:
            result = analyze_transcript_pdf(temp_path, max_pages=20)
            return jsonify(result)
        finally:
            try:
                os.remove(temp_path)
            except OSError:
                pass

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/fraud/transcript/report', methods=['POST'])
def export_transcript_fraud_report():
    """Export one transcript fraud report as PDF or TXT."""
    redirect_response = require_login()
    if redirect_response:
        return redirect_response

    try:
        payload = request.get_json(silent=True) or {}
        result = payload.get("result")

        if not isinstance(result, dict):
            return jsonify({"error": "Missing or invalid 'result' object"}), 400

        comments = payload.get("comments", "") or ""
        meta = payload.get("meta", {}) or {}
        fmt = (payload.get("format", "pdf") or "pdf").lower()
        original_filename = payload.get("originalFilename", "") or ""

        reviewer = session.get("full_name") or session.get("username") or "unknown"

        report_text = build_transcript_txt_report(
            result=result,
            reviewer=reviewer,
            comments=comments,
            meta=meta
        )

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = meta.get("studentId") or os.path.splitext(original_filename)[0] or "transcript_fraud"
        safe_base = re.sub(r"[^A-Za-z0-9_\-]+", "_", base).strip("_") or "transcript_fraud"

        if fmt == "txt":
            return send_file(
                io.BytesIO(report_text.encode("utf-8")),
                mimetype="text/plain",
                as_attachment=True,
                download_name=f"{safe_base}_report_{ts}.txt"
            )

        pdf_bytes = build_pdf_bytes_from_text(report_text)
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"{safe_base}_report_{ts}.pdf"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/fraud/transcript/report/all', methods=['POST'])
def export_transcript_fraud_report_all():
    """Export one combined transcript fraud report for multiple documents."""
    redirect_response = require_login()
    if redirect_response:
        return redirect_response

    try:
        payload = request.get_json(silent=True) or {}
        documents = payload.get("documents", [])
        fmt = (payload.get("format", "pdf") or "pdf").lower()

        if not isinstance(documents, list) or len(documents) == 0:
            return jsonify({"error": "No documents provided for export."}), 400

        cleaned = []
        for d in documents:
            if not isinstance(d, dict):
                continue
            r = d.get("result")
            if not isinstance(r, dict):
                continue
            cleaned.append({
                "originalFilename": d.get("originalFilename", "") or "",
                "result": r,
                "comments": d.get("comments", "") or "",
                "meta": d.get("meta", {}) or {}
            })

        if len(cleaned) == 0:
            return jsonify({"error": "No valid document results found for export."}), 400

        reviewer = session.get("full_name") or session.get("username") or "unknown"
        report_text = build_multi_transcript_txt_report(cleaned, reviewer=reviewer)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_base = "transcript_fraud_combined"

        if fmt == "txt":
            return send_file(
                io.BytesIO(report_text.encode("utf-8")),
                mimetype="text/plain",
                as_attachment=True,
                download_name=f"{safe_base}_{ts}.txt"
            )

        pdf_bytes = build_pdf_bytes_from_text(report_text)
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"{safe_base}_{ts}.pdf"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

def build_financial_txt_report(
    result: dict,
    reviewer: str,
    comments: str = "",
    meta: Optional[Dict] = None,
    include_header: bool = True
) -> str:

    meta = meta or {}
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")

    file_path = result.get("file_path", "")
    doc_sev = result.get("doc_severity", "UNKNOWN")
    pages = result.get("pages", []) or []

    def safe(s: str, width: int = 240) -> str:
        return shorten((s or "").strip().replace("\n", " "), width=width, placeholder="...")

    lines = []

    if include_header:
        lines.append("INTERNATIONAL APPLICANT RATING ALGORITHM (IARA)")
        lines.append("Financial Document Fraud Screening Report")
        lines.append("=" * 78)
        lines.append(f"Generated: {now}")
        lines.append(f"Reviewer: {reviewer}")

        if meta.get("studentId"):
            lines.append(f"Student ID: {meta.get('studentId')}")
        if meta.get("applicantName"):
            lines.append(f"Applicant Name: {meta.get('applicantName')}")
        if meta.get("program"):
            lines.append(f"Program/Term: {meta.get('program')}")
        lines.append("-" * 78)


    lines.append(f"Document: {os.path.basename(file_path) if file_path else '(uploaded file)'}")
    lines.append(f"Document Severity: {doc_sev}")
    lines.append(f"Pages Analyzed: {len(pages)}")
    lines.append("-" * 78)

    # Only include fraud-positive signals in the justification section (clean & defensible)
    for p in pages:
        pnum = p.get("page_number")
        sev = p.get("severity", "UNKNOWN")
        conf = float(p.get("confidence", 0.0) or 0.0)

        summary = safe(p.get("ai_summary", ""), width=260)
        lines.append(f"Page {pnum}: Severity={sev}  Confidence={conf:.2f}")
        if summary:
            lines.append(f"  Summary: {summary}")

        sigs = p.get("fraud_signals", []) or []
        fraud_pos = [
            s for s in sigs
            if isinstance(s, dict) and (s.get("polarity") == "fraud_positive")
        ]

        if fraud_pos:
            lines.append("  Fraud-Positive Signals:")
            for s in fraud_pos[:12]:
                signal = safe(s.get("signal", ""), width=260)
                cat = s.get("category", "other")
                src = s.get("source", "llm")
                sc = float(s.get("confidence", 0.0) or 0.0)
                lines.append(f"    - [{cat}] ({src}, {sc:.2f}) {signal}")
        else:
            lines.append("  Fraud-Positive Signals: None detected on this page.")

        lines.append("")

    lines.append("-" * 78)
    lines.append("Reviewer Comments (for applicant file):")
    lines.append((comments or "").strip() if (comments or "").strip() else "(none)")
    lines.append("=" * 78)
    lines.append("Note: This report is a pre-screening aid. Final decisions require human review.")
    return "\n".join(lines)

def build_multi_financial_txt_report(documents: list, reviewer: str) -> str:
    """
    documents: list of dicts like:
      {
        "originalFilename": "...pdf",
        "result": { ... detector JSON ... },
        "comments": "...",
        "meta": { "studentId": "...", "applicantName": "...", "program": "..." }
      }
    """
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    lines = []
    lines.append("INTERNATIONAL APPLICANT RATING ALGORITHM (IARA)")
    lines.append("Financial Document Fraud Screening Report (Combined)")
    lines.append("=" * 78)
    lines.append(f"Generated: {now}")
    lines.append(f"Reviewer: {reviewer}")
    lines.append(f"Documents Included: {len(documents)}")
    lines.append("=" * 78)
    lines.append("")

    for idx, doc in enumerate(documents, start=1):
        result = doc.get("result") or {}
        comments = doc.get("comments") or ""
        meta = doc.get("meta") or {}
        original_filename = doc.get("originalFilename") or "(unknown file)"

        # Section header for each PDF
        lines.append("#" * 78)
        lines.append(f"DOCUMENT {idx}: {original_filename}")
        if meta.get("studentId"):
            lines.append(f"Student ID: {meta.get('studentId')}")
        if meta.get("applicantName"):
            lines.append(f"Applicant Name: {meta.get('applicantName')}")
        if meta.get("program"):
            lines.append(f"Program/Term: {meta.get('program')}")
        lines.append("#" * 78)

        # Reuse your single-doc builder (keeps style consistent)
        lines.append(build_financial_txt_report(
            result=result,
            reviewer=reviewer,
            comments=comments,
            meta=meta,
            include_header=False
        ))

        lines.append("")  # spacer

    return "\n".join(lines)

def build_transcript_txt_report(
    result: dict,
    reviewer: str,
    comments: str = "",
    meta: Optional[Dict] = None,
    include_header: bool = True
) -> str:
    meta = meta or {}
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")

    file_path = result.get("file_path", "")
    doc_sev = result.get("doc_severity", "UNKNOWN")
    pages = result.get("pages", []) or []

    def safe(s: str, width: int = 240) -> str:
        return shorten((s or "").strip().replace("\n", " "), width=width, placeholder="...")

    lines = []

    if include_header:
        lines.append("INTERNATIONAL APPLICANT RATING ALGORITHM (IARA)")
        lines.append("Transcript Fraud Pre-Screening Report")
        lines.append("=" * 78)
        lines.append(f"Generated: {now}")
        lines.append(f"Reviewer: {reviewer}")

        if meta.get("studentId"):
            lines.append(f"Student ID: {meta.get('studentId')}")
        if meta.get("applicantName"):
            lines.append(f"Applicant Name: {meta.get('applicantName')}")
        if meta.get("program"):
            lines.append(f"Program/Term: {meta.get('program')}")
        lines.append("-" * 78)

    lines.append(f"Document: {os.path.basename(file_path) if file_path else '(uploaded file)'}")
    lines.append(f"Document Severity: {doc_sev}")
    lines.append(f"Pages Analyzed: {len(pages)}")
    lines.append("-" * 78)

    # Only include fraud-positive signals in justification section (defensible)
    for p in pages:
        pnum = p.get("page_number")
        sev = p.get("severity", "UNKNOWN")
        conf = float(p.get("confidence", 0.0) or 0.0)

        summary = safe(p.get("ai_summary", ""), width=260)
        lines.append(f"Page {pnum}: Severity={sev}  Confidence={conf:.2f}")
        if summary:
            lines.append(f"  Summary: {summary}")

        sigs = p.get("fraud_signals", []) or []
        fraud_pos = [
            s for s in sigs
            if isinstance(s, dict) and (s.get("polarity") == "fraud_positive")
        ]

        if fraud_pos:
            lines.append("  Fraud-Positive Signals:")
            for s in fraud_pos[:12]:
                signal = safe(s.get("signal", ""), width=260)
                cat = s.get("category", "other")
                src = s.get("source", "llm")
                sc = float(s.get("confidence", 0.0) or 0.0)
                lines.append(f"    - [{cat}] ({src}, {sc:.2f}) {signal}")
        else:
            lines.append("  Fraud-Positive Signals: None detected on this page.")

        lines.append("")

    lines.append("-" * 78)
    lines.append("Reviewer Comments (for applicant file):")
    lines.append((comments or "").strip() if (comments or "").strip() else "(none)")
    lines.append("=" * 78)
    lines.append("Note: This report is a pre-screening aid. Final decisions require human review.")
    return "\n".join(lines)

def build_multi_transcript_txt_report(documents: list, reviewer: str) -> str:
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    lines = []
    lines.append("INTERNATIONAL APPLICANT RATING ALGORITHM (IARA)")
    lines.append("Transcript Fraud Pre-Screening Report (Combined)")
    lines.append("=" * 78)
    lines.append(f"Generated: {now}")
    lines.append(f"Reviewer: {reviewer}")
    lines.append(f"Documents Included: {len(documents)}")
    lines.append("=" * 78)
    lines.append("")

    for idx, doc in enumerate(documents, start=1):
        result = doc.get("result") or {}
        comments = doc.get("comments") or ""
        meta = doc.get("meta") or {}
        original_filename = doc.get("originalFilename") or "(unknown file)"

        lines.append("#" * 78)
        lines.append(f"DOCUMENT {idx}: {original_filename}")
        if meta.get("studentId"):
            lines.append(f"Student ID: {meta.get('studentId')}")
        if meta.get("applicantName"):
            lines.append(f"Applicant Name: {meta.get('applicantName')}")
        if meta.get("program"):
            lines.append(f"Program/Term: {meta.get('program')}")
        lines.append("#" * 78)

        lines.append(build_transcript_txt_report(
            result=result,
            reviewer=reviewer,
            comments=comments,
            meta=meta,
            include_header=False
        ))

        lines.append("")

    return "\n".join(lines)

def wrap_line_for_pdf(line: str, canvas_obj, max_width: float, font_name: str, font_size: int):
    if not line:
        return [""]
    
    words = line.split(" ")
    wrapped_lines = []
    current_line = ""
    
    for word in words:
        test_line = word if not current_line else current_line + " " + word
        
        if canvas_obj.stringWidth(test_line, font_name, font_size) <= max_width:
            current_line = test_line
        else:
            if current_line:
                wrapped_lines.append(current_line)
            current_line = ""
            
            # If a single word is too long (like a filename), break it into chunks
            if canvas_obj.stringWidth(word, font_name, font_size) > max_width:
                chunk = ""
                for ch in word:
                    test_chunk = chunk + ch
                    if canvas_obj.stringWidth(test_chunk, font_name, font_size) <= max_width:
                        chunk = test_chunk
                    else:
                        if chunk:
                            wrapped_lines.append(chunk)
                        chunk = ch
                if chunk:
                    current_line = chunk
            else:
                current_line = word
    
    if current_line:
        wrapped_lines.append(current_line)
    
    return wrapped_lines

def build_pdf_bytes_from_text(report_text: str) -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    width, height = letter

    x = 0.75 * inch
    y = height - 0.75 * inch
    right_margin = 0.75 * inch
    max_width = width - x - right_margin

    font_name = "Courier"
    font_size = 10
    line_height = 12

    c.setFont(font_name, font_size)

    for raw_line in report_text.splitlines():
        line = raw_line or ""

        wrapped_lines = wrap_line_for_pdf(line, c, max_width, font_name, font_size)
        if not wrapped_lines:
            wrapped_lines = [""]

        for wrapped_line in wrapped_lines:
            if y < 0.75 * inch:
                c.showPage()
                c.setFont(font_name, font_size)
                y = height - 0.75 * inch

            c.drawString(x, y, wrapped_line)
            y -= line_height

    c.save()
    buf.seek(0)
    return buf.getvalue()


if __name__ == '__main__':
    app.run(debug=True, port=5000)